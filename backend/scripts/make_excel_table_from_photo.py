#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("outputs/afm_photo_table_template.xlsx")


PERIOD_2022_2026_COLUMNS = [
    "Наименование ИП / ИП",
    "ИИН/БИН",
    "Дата",
    "Общая сумма",
    "Ошибки по СМР",
    "Услуги",
    "Работы",
    "Договор / работа / общая сумма",
    "ФЛ / ИП / юр. лица",
    "Госзакуп / тендер",
    "Часто встречающиеся суммы",
    "Риски",
]


FIRST_HALF_COLUMNS = [
    "Наименование ИП / ИП",
    "БОКАТУ / регион",
    "ФЛ / ИП",
    "Общая сумма",
    "Ошибки по СМР",
    "Услуги",
    "Договор / работа / общая сумма",
    "Госзакуп / тендер",
    "Оценка / вывод",
    "Риски",
]


SAMPLE_ROWS_PERIOD = [
    ["ИП Иванов", "", "", "", "", "", "", "", "", "", "", ""],
    ["Проведение мероприятий", "", "", "", "", "", "", "", "", "", "", ""],
    ["Оказание услуг", "", "", "", "", "", "", "", "", "", "", ""],
    ["Общие бюджетные риски", "", "", "", "", "", "", "", "", "", "", ""],
]


SAMPLE_ROWS_HALF = [
    ["ИП Иванов", "", "", "", "", "", "", "", "", ""],
    ["Проблемные операции", "", "", "", "", "", "", "", "", ""],
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel template based on the photographed hand-drawn table."
    )
    parser.add_argument(
        "-o",
        "--output",
        default=str(OUTPUT_PATH),
        help=f"Output xlsx path. Default: {OUTPUT_PATH}",
    )
    return parser.parse_args()


def write_row(ws, row: int, values: Iterable[str], *, fill: PatternFill, font: Font) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_table_style(ws, min_row: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color="4F5B66")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def add_section(
    ws,
    *,
    start_row: int,
    title: str,
    subtitle: str,
    columns: list[str],
    sample_rows: list[list[str]],
    note: str,
) -> int:
    max_col = len(columns)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True, size=10)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    title_cell = ws.cell(start_row, 1, title)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=max_col)
    subtitle_cell = ws.cell(start_row + 1, 1, subtitle)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    subtitle_cell.font = Font(italic=True, color="404040")

    header_row = start_row + 3
    write_row(ws, header_row, columns, fill=header_fill, font=header_font)

    for offset, row_values in enumerate(sample_rows, start=1):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=header_row + offset, column=col, value=value)

    total_row = header_row + len(sample_rows) + 2
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=max_col)
    total_cell = ws.cell(total_row, 1, note)
    total_cell.font = Font(bold=True, italic=True)
    total_cell.alignment = Alignment(horizontal="center")

    apply_table_style(ws, header_row, total_row, max_col)
    ws.row_dimensions[start_row].height = 26
    ws.row_dimensions[start_row + 1].height = 24
    ws.row_dimensions[header_row].height = 48

    return total_row + 3


def create_workbook(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон анализа"
    ws.freeze_panes = "A5"

    next_row = add_section(
        ws,
        start_row=1,
        title="Период 2022-2026",
        subtitle="Интеграция: ИП -> общий анализ",
        columns=PERIOD_2022_2026_COLUMNS,
        sample_rows=SAMPLE_ROWS_PERIOD,
        note="Итого / ориентир: 90-220 млрд",
    )

    add_section(
        ws,
        start_row=next_row,
        title="Первое полугодие 2022-2026",
        subtitle="Отдельный блок для сверки показателей за первое полугодие",
        columns=FIRST_HALF_COLUMNS,
        sample_rows=SAMPLE_ROWS_HALF,
        note="Итого / ориентир: 100 млрд",
    )

    ws.sheet_view.showGridLines = False
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    output_path = Path(args.output)
    create_workbook(output_path)
    print(f"Excel file created: {output_path.resolve()}")


if __name__ == "__main__":
    main()
